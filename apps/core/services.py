from django.db import models
from django.db.models import Q, Count
from django.utils import timezone
from datetime import datetime, timedelta
from decimal import Decimal
from .models import User
from .models_audit import AuditLog


class UserService:
    @staticmethod
    def list_users(role=None, is_active=None, search=None):
        queryset = User.objects.all()
        if role:
            queryset = queryset.filter(role=role)
        if is_active is not None:
            queryset = queryset.filter(is_active=is_active)
        if search:
            queryset = queryset.filter(
                Q(username__icontains=search) |
                Q(email__icontains=search) |
                Q(phone__icontains=search)
            )
        return queryset.select_related()

    @staticmethod
    def get_coaches():
        return User.objects.filter(role=User.Role.COACH, is_active=True)

    @staticmethod
    def create_user(**kwargs):
        password = kwargs.pop('password', None)
        user = User(**kwargs)
        if password:
            user.set_password(password)
        user.save()
        return user


class AuditService:
    @staticmethod
    def _serialize_value(value):
        import datetime as dt
        if value is None:
            return None
        if isinstance(value, (int, float, str, bool)):
            return value
        if isinstance(value, (dt.datetime, dt.date)):
            return value.isoformat()
        if isinstance(value, Decimal):
            return str(value)
        if hasattr(value, 'pk'):
            return str(value.pk)
        if hasattr(value, '__str__'):
            return str(value)
        return str(value)

    @staticmethod
    def _serialize_dict(data):
        if not data:
            return data
        result = {}
        for key, value in data.items():
            result[key] = AuditService._serialize_value(value)
        return result

    @staticmethod
    def log_action(user, action, instance=None, old_value=None, new_value=None, ip_address=None, user_agent=None):
        from django.contrib.contenttypes.models import ContentType
        content_type = None
        object_id = None
        object_repr = str(instance) if instance else 'N/A'

        if instance and hasattr(instance, 'pk'):
            content_type = ContentType.objects.get_for_model(instance)
            object_id = instance.pk

        old_value = AuditService._serialize_dict(old_value)
        new_value = AuditService._serialize_dict(new_value)

        return AuditLog.objects.create(
            user=user,
            action=action,
            content_type=content_type,
            object_id=object_id,
            object_repr=object_repr,
            old_value=old_value,
            new_value=new_value,
            ip_address=ip_address,
            user_agent=user_agent or '',
        )

    @staticmethod
    def list_logs(user=None, action=None, start_date=None, end_date=None, content_type=None, object_id=None, search=None):
        queryset = AuditLog.objects.all()
        if user:
            queryset = queryset.filter(user=user)
        if action:
            queryset = queryset.filter(action=action)
        if start_date:
            queryset = queryset.filter(created_at__gte=start_date)
        if end_date:
            queryset = queryset.filter(created_at__lte=end_date + timedelta(days=1))
        if content_type:
            queryset = queryset.filter(content_type__model=content_type)
        if object_id:
            queryset = queryset.filter(object_id=object_id)
        if search:
            queryset = queryset.filter(
                Q(object_repr__icontains=search) |
                Q(user__username__icontains=search)
            )
        return queryset.select_related('user', 'content_type')

    @staticmethod
    def get_trail_for_instance(content_type, object_id):
        return AuditLog.objects.filter(
            content_type__model=content_type,
            object_id=object_id
        ).order_by('-created_at').select_related('user')


class ExportService:
    @staticmethod
    def export_to_excel(queryset, fields, filename_prefix='export'):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from django.http import HttpResponse
        import datetime

        wb = Workbook()
        ws = wb.active
        ws.title = '数据导出'

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')

        for col, field in enumerate(fields, 1):
            cell = ws.cell(row=1, column=col, value=field['label'])
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            ws.column_dimensions[chr(64 + col)].width = field.get('width', 15)

        for row, obj in enumerate(queryset, 2):
            for col, field in enumerate(fields, 1):
                value = field['value'](obj) if callable(field['value']) else getattr(obj, field['value'], '')
                if isinstance(value, datetime.datetime):
                    value = value.strftime('%Y-%m-%d %H:%M:%S')
                elif isinstance(value, datetime.date):
                    value = value.strftime('%Y-%m-%d')
                ws.cell(row=row, column=col, value=str(value))

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename_prefix}_{timezone.now().strftime("%Y%m%d%H%M%S")}.xlsx"'
        wb.save(response)
        return response
