from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.utils import timezone
from ..models import (
    Store, Product, Inventory,
    ReplenishmentOrder, TransferOrder, DisplayRecord, MemberRedemption
)
import io
import logging

logger = logging.getLogger(__name__)


class ExportService:
    """导出服务 - 支持多格式导出"""

    EXCEL_CONTENT_TYPE = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    @staticmethod
    def _create_workbook_with_header(sheet_name, headers, column_widths=None):
        """创建带表头的工作簿"""
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            if column_widths and col - 1 < len(column_widths):
                ws.column_dimensions[get_column_letter(col)].width = column_widths[col - 1]

        return wb, ws

    @staticmethod
    def _add_data_rows(ws, data_rows, start_row=2):
        """添加数据行"""
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        for row_idx, row_data in enumerate(data_rows, start_row):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center')
        return ws

    @staticmethod
    def _create_response(wb, filename):
        """创建HTTP响应"""
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(buffer.read(), content_type=ExportService.EXCEL_CONTENT_TYPE)
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    @classmethod
    def export_replenishment_orders(cls, queryset, user, request=None):
        """导出货单"""
        from .audit_service import AuditService

        headers = [
            '补货单号', '门店', '状态', '优先级', '关联计划',
            '商品SKU', '商品名称', '申请数量', '批准数量', '实发数量',
            '实收数量', '单价', '金额', '申请备注',
            '创建时间', '提交时间', '审核时间', '发货时间', '收货时间'
        ]
        widths = [18, 20, 10, 8, 20, 15, 25, 10, 10, 10, 10, 10, 12, 20, 20, 20, 20, 20, 20]

        wb, ws = cls._create_workbook_with_header('补货单明细', headers, widths)

        rows = []
        for order in queryset:
            for item in order.items.all():
                amount = (item.received_quantity or item.shipped_quantity or
                          item.approved_quantity or item.requested_quantity) * item.unit_price
                rows.append([
                    order.code,
                    order.store.name,
                    order.get_status_display(),
                    order.get_priority_display(),
                    order.plan.name if order.plan else '',
                    item.product.sku,
                    item.product.name,
                    item.requested_quantity,
                    item.approved_quantity or '',
                    item.shipped_quantity or '',
                    item.received_quantity or '',
                    float(item.unit_price),
                    float(amount),
                    order.remark,
                    order.created_at.strftime('%Y-%m-%d %H:%M') if order.created_at else '',
                    order.submitted_at.strftime('%Y-%m-%d %H:%M') if order.submitted_at else '',
                    order.reviewed_at.strftime('%Y-%m-%d %H:%M') if order.reviewed_at else '',
                    order.shipped_at.strftime('%Y-%m-%d %H:%M') if order.shipped_at else '',
                    order.received_at.strftime('%Y-%m-%d %H:%M') if order.received_at else '',
                ])

        cls._add_data_rows(ws, rows)
        AuditService.log_export(user, 'ReplenishmentOrder', str(queryset.query), len(rows), request)
        filename = f'补货单_{timezone.localdate().strftime("%Y%m%d")}.xlsx'
        return cls._create_response(wb, filename)

    @classmethod
    def export_transfer_orders(cls, queryset, user, request=None):
        """导出调拨单"""
        from .audit_service import AuditService

        headers = [
            '调拨单号', '转出门店', '转入门店', '状态', '调拨原因',
            '商品SKU', '商品名称', '调拨数量', '实际转出', '实际转入',
            '创建时间', '提交时间', '转出确认时间', '转入确认时间'
        ]
        widths = [18, 20, 20, 10, 25, 15, 25, 10, 10, 10, 20, 20, 20, 20]

        wb, ws = cls._create_workbook_with_header('调拨单明细', headers, widths)

        rows = []
        for order in queryset:
            for item in order.items.all():
                rows.append([
                    order.code,
                    order.from_store.name,
                    order.to_store.name,
                    order.get_status_display(),
                    order.reason,
                    item.product.sku,
                    item.product.name,
                    item.transfer_quantity,
                    item.out_quantity or '',
                    item.in_quantity or '',
                    order.created_at.strftime('%Y-%m-%d %H:%M') if order.created_at else '',
                    order.submitted_at.strftime('%Y-%m-%d %H:%M') if order.submitted_at else '',
                    order.out_confirmed_at.strftime('%Y-%m-%d %H:%M') if order.out_confirmed_at else '',
                    order.in_confirmed_at.strftime('%Y-%m-%d %H:%M') if order.in_confirmed_at else '',
                ])

        cls._add_data_rows(ws, rows)
        AuditService.log_export(user, 'TransferOrder', str(queryset.query), len(rows), request)
        filename = f'调拨单_{timezone.localdate().strftime("%Y%m%d")}.xlsx'
        return cls._create_response(wb, filename)

    @classmethod
    def export_inventory(cls, queryset, user, request=None):
        """导出库存"""
        from .audit_service import AuditService

        headers = [
            '门店编码', '门店名称', 'SKU编码', '商品名称', '品类',
            '可用库存', '预留库存', '安全库存', '库存状态', '最近盘点时间'
        ]
        widths = [12, 20, 15, 25, 15, 10, 10, 10, 12, 20]

        wb, ws = cls._create_workbook_with_header('库存明细', headers, widths)

        rows = []
        for inv in queryset:
            if inv.quantity <= 0:
                status = '缺货'
            elif inv.quantity <= inv.product.safe_stock:
                status = '低于安全库存'
            else:
                status = '正常'
            rows.append([
                inv.store.code,
                inv.store.name,
                inv.product.sku,
                inv.product.name,
                inv.product.category,
                inv.quantity,
                inv.reserved_quantity,
                inv.product.safe_stock,
                status,
                inv.last_counted_at.strftime('%Y-%m-%d %H:%M') if inv.last_counted_at else '未盘点',
            ])

        cls._add_data_rows(ws, rows)
        AuditService.log_export(user, 'Inventory', str(queryset.query), len(rows), request)
        filename = f'库存_{timezone.localdate().strftime("%Y%m%d")}.xlsx'
        return cls._create_response(wb, filename)

    @classmethod
    def export_display_records(cls, queryset, user, request=None):
        """导出陈列记录"""
        from .audit_service import AuditService

        headers = [
            '门店', '商品SKU', '商品名称', '检查日期', '问题类型',
            '问题描述', '状态', '整改说明', '检查人', '整改人', '复核人'
        ]
        widths = [20, 15, 25, 12, 15, 30, 10, 30, 12, 12, 12]

        wb, ws = cls._create_workbook_with_header('陈列记录', headers, widths)

        rows = []
        for record in queryset:
            rows.append([
                record.store.name,
                record.product.sku,
                record.product.name,
                record.check_date.strftime('%Y-%m-%d'),
                record.issue_type,
                record.description,
                record.get_status_display(),
                record.fix_note,
                record.checked_by.username if record.checked_by else '',
                record.fixed_by.username if record.fixed_by else '',
                record.verified_by.username if record.verified_by else '',
            ])

        cls._add_data_rows(ws, rows)
        AuditService.log_export(user, 'DisplayRecord', str(queryset.query), len(rows), request)
        filename = f'陈列记录_{timezone.localdate().strftime("%Y%m%d")}.xlsx'
        return cls._create_response(wb, filename)

    @classmethod
    def export_redemptions(cls, queryset, user, request=None):
        """导出会员兑换"""
        from .audit_service import AuditService

        headers = [
            '兑换单号', '会员姓名', '会员电话', '门店', '商品SKU',
            '商品名称', '兑换数量', '消耗积分', '状态', '创建时间'
        ]
        widths = [18, 12, 15, 20, 15, 25, 10, 10, 10, 20]

        wb, ws = cls._create_workbook_with_header('会员兑换', headers, widths)

        rows = []
        for redemption in queryset:
            rows.append([
                redemption.code,
                redemption.member_name,
                redemption.member_phone,
                redemption.store.name,
                redemption.product.sku,
                redemption.product.name,
                redemption.quantity,
                redemption.points_used,
                redemption.get_status_display(),
                redemption.created_at.strftime('%Y-%m-%d %H:%M'),
            ])

        cls._add_data_rows(ws, rows)
        AuditService.log_export(user, 'MemberRedemption', str(queryset.query), len(rows), request)
        filename = f'会员兑换_{timezone.localdate().strftime("%Y%m%d")}.xlsx'
        return cls._create_response(wb, filename)
